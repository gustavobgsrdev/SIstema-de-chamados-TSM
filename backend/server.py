from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import base64
from io import BytesIO
from PIL import Image
import asyncio

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'service_order_db')]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# ============ MODELS ============

class UserCreate(BaseModel):
    email: str  # Username
    password: str
    name: str
    role: str = "USER"

class UserLogin(BaseModel):
    email: str  # Username
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str  # Username (not email format)
    name: str
    role: str = "USER"  # ADMIN or USER
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User

class EquipmentInfo(BaseModel):
    serial_number: Optional[str] = None
    board_serial: Optional[str] = None

class Verification(BaseModel):
    item: str
    status: str  # BOA, RUIM, N/A
    observation: Optional[str] = None

class ServiceOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    
    # Informações básicas
    ticket_number: Optional[str] = None
    os_number: Optional[str] = None
    pat: Optional[str] = None
    status: Optional[str] = "ABERTO"  # URGENTE, ABERTO, EM ROTA, LIBERADO, PENDENCIA, SUSPENSO, DEFINIR, RESOLVIDO
    opening_date: Optional[str] = None
    responsible_opening: Optional[str] = None
    responsible_tech: Optional[str] = None
    phone: Optional[str] = None
    
    # Cliente
    client_name: Optional[str] = None
    unit: Optional[str] = None
    service_address: Optional[str] = None
    
    # Equipamento
    equipment_type: Optional[str] = None  # Ex: IMPRESSORA
    equipment_brand: Optional[str] = None  # Ex: SAMSUNG
    equipment_model: Optional[str] = None  # Ex: M4070FR
    equipment_serial: Optional[str] = None
    equipment_board_serial: Optional[str] = None
    
    # Chamado
    call_info: Optional[str] = None
    materials: Optional[str] = None
    technical_report: Optional[str] = None
    
    # Verificações
    verifications: List[Verification] = Field(default_factory=list)
    
    # Contadores e pendências
    total_page_count: Optional[str] = None
    pending_issues: Optional[str] = None
    next_visit: Optional[str] = None
    equipment_replaced: Optional[bool] = False
    
    # Observações
    observations: Optional[str] = None
    
    # Metadata
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ServiceOrderCreate(BaseModel):
    ticket_number: Optional[str] = None
    os_number: Optional[str] = None
    pat: Optional[str] = None
    status: Optional[str] = "ABERTO"
    opening_date: Optional[str] = None
    responsible_opening: Optional[str] = None
    responsible_tech: Optional[str] = None
    phone: Optional[str] = None
    client_name: Optional[str] = None
    unit: Optional[str] = None
    service_address: Optional[str] = None
    equipment_type: Optional[str] = None
    equipment_brand: Optional[str] = None
    equipment_model: Optional[str] = None
    equipment_serial: Optional[str] = None
    equipment_board_serial: Optional[str] = None
    call_info: Optional[str] = None
    materials: Optional[str] = None
    technical_report: Optional[str] = None
    verifications: List[Verification] = Field(default_factory=list)
    total_page_count: Optional[str] = None
    pending_issues: Optional[str] = None
    next_visit: Optional[str] = None
    equipment_replaced: Optional[bool] = False
    observations: Optional[str] = None

class ServiceOrderUpdate(BaseModel):
    ticket_number: Optional[str] = None
    os_number: Optional[str] = None
    pat: Optional[str] = None
    status: Optional[str] = None
    opening_date: Optional[str] = None
    responsible_opening: Optional[str] = None
    responsible_tech: Optional[str] = None
    phone: Optional[str] = None
    client_name: Optional[str] = None
    unit: Optional[str] = None
    service_address: Optional[str] = None
    equipment_type: Optional[str] = None
    equipment_brand: Optional[str] = None
    equipment_model: Optional[str] = None
    equipment_serial: Optional[str] = None
    equipment_board_serial: Optional[str] = None
    call_info: Optional[str] = None
    materials: Optional[str] = None
    technical_report: Optional[str] = None
    verifications: Optional[List[Verification]] = None
    total_page_count: Optional[str] = None
    pending_issues: Optional[str] = None
    next_visit: Optional[str] = None
    equipment_replaced: Optional[bool] = None
    observations: Optional[str] = None

class OCRResponse(BaseModel):
    extracted_text: str
    structured_data: dict

# ============ AUTH FUNCTIONS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(user_id: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "user_id": user_id,
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user_doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")
        
        # Convert ISO string to datetime if needed
        if isinstance(user_doc.get('created_at'), str):
            user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
        
        return User(**user_doc)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ OCR FUNCTION ============

async def extract_text_from_image(image_base64: str) -> dict:
    """Extract text from image and structure it for service order form"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        import json
        
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="EMERGENT_LLM_KEY not configured")
        
        # Create chat instance
        chat = LlmChat(
            api_key=api_key,
            session_id=str(uuid.uuid4()),
            system_message="""You are an intelligent OCR assistant specialized in extracting service order information from images.
Extract all visible information and return it in a structured JSON format that matches these exact field names:
{
  "ticket_number": "número do chamado",
  "os_number": "número da O.S.",
  "pat": "código PAT se presente",
  "opening_date": "data de abertura (formato: DD/MM/YYYY ou DD/MM)",
  "responsible_opening": "nome do responsável pela abertura",
  "responsible_tech": "nome do técnico responsável",
  "phone": "telefone",
  "client_name": "nome do cliente",
  "unit": "unidade/local",
  "service_address": "endereço completo de atendimento",
  "equipment_serial": "número de série do equipamento (S/N EQUIP)",
  "equipment_board_serial": "número de série da placa (S/N PLACA)",
  "equipment_type": "tipo de equipamento (IMPRESSORA, etc)",
  "equipment_brand": "marca do equipamento",
  "equipment_model": "modelo do equipamento",
  "call_info": "descrição do problema/chamado",
  "materials": "materiais utilizados",
  "technical_report": "laudo técnico",
  "total_page_count": "contador de páginas",
  "observations": "observações gerais"
}
Return ONLY valid JSON. If a field is not found, use null."""
        ).with_model("openai", "gpt-5.1")
        
        # Create image content
        image_content = ImageContent(image_base64=image_base64)
        
        # Create message
        user_message = UserMessage(
            text="""Analyze this service order image and extract ALL visible information. 
Pay special attention to:
- Números (chamado, O.S., PAT)
- Datas (abertura, próxima visita)
- Nomes (responsáveis, técnico, cliente)
- Equipamento (marca, modelo, números de série da placa e equipamento)
- Endereço completo
- Descrição do problema
- Materiais utilizados

Return the data in the exact JSON format specified in the system message. Be precise and extract everything visible.""",
            file_contents=[image_content]
        )
        
        # Send message and get response
        response = await chat.send_message(user_message)
        
        # Try to parse JSON from response
        try:
            # Clean response - remove markdown code blocks if present
            cleaned_response = response.strip()
            if cleaned_response.startswith("```"):
                # Remove ```json or ``` from start
                cleaned_response = cleaned_response.split("\\n", 1)[1]
            if cleaned_response.endswith("```"):
                # Remove ``` from end
                cleaned_response = cleaned_response.rsplit("\\n", 1)[0]
            
            structured_data = json.loads(cleaned_response)
        except json.JSONDecodeError:
            # If JSON parsing fails, return raw text
            logging.warning(f"Could not parse JSON from OCR response: {response}")
            structured_data = {
                "raw_ocr": response,
                "parse_error": True
            }
        
        return {
            "extracted_text": response,
            "structured_data": structured_data
        }
    except Exception as e:
        logging.error(f"OCR Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

# ============ ROUTES ============

@api_router.get("/")
async def root():
    return {"message": "Service Order Management API"}

# Auth routes
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    # Only ADMIN can create new users
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Only administrators can create new users")
    
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role
    )
    
    user_doc = user.model_dump()
    user_doc['password'] = hash_password(user_data.password)
    user_doc['created_at'] = user_doc['created_at'].isoformat()
    
    await db.users.insert_one(user_doc)
    
    # Create token
    token = create_access_token(user.id)
    
    return TokenResponse(access_token=token, user=user)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email})
    
    if not user_doc or not verify_password(credentials.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Convert ISO string to datetime
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    user = User(**{k: v for k, v in user_doc.items() if k != 'password'})
    token = create_access_token(user.id)
    
    return TokenResponse(access_token=token, user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# User management routes (ADMIN only)
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Only administrators can view users")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    # Convert ISO strings to datetime
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return users

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Only administrators can delete users")
    
    # Cannot delete yourself
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# Service Order routes
@api_router.post("/service-orders", response_model=ServiceOrder)
async def create_service_order(
    order_data: ServiceOrderCreate,
    current_user: User = Depends(get_current_user)
):
    order = ServiceOrder(
        **order_data.model_dump(),
        created_by=current_user.id
    )
    
    order_doc = order.model_dump()
    order_doc['created_at'] = order_doc['created_at'].isoformat()
    order_doc['updated_at'] = order_doc['updated_at'].isoformat()
    
    await db.service_orders.insert_one(order_doc)
    
    return order

@api_router.get("/service-orders", response_model=List[ServiceOrder])
async def get_service_orders(
    current_user: User = Depends(get_current_user),
    status: Optional[str] = None,
    pat: Optional[str] = None,
    ticket_number: Optional[str] = None,
    os_number: Optional[str] = None,
    equipment_serial: Optional[str] = None,
    unit: Optional[str] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None
):
    # Build filter
    filter_query = {}
    if status and status.strip():
        filter_query['status'] = status
    if pat:
        filter_query['pat'] = {"$regex": pat, "$options": "i"}
    if ticket_number:
        filter_query['ticket_number'] = {"$regex": ticket_number, "$options": "i"}
    if os_number:
        filter_query['os_number'] = {"$regex": os_number, "$options": "i"}
    if equipment_serial:
        filter_query['equipment_serial'] = {"$regex": equipment_serial, "$options": "i"}
    if unit:
        filter_query['unit'] = {"$regex": unit, "$options": "i"}
    
    # Date range filter
    if date_start or date_end:
        date_filter = {}
        if date_start:
            date_filter["$gte"] = date_start
        if date_end:
            date_filter["$lte"] = date_end
        if date_filter:
            filter_query['opening_date'] = date_filter
    
    # Get orders sorted by creation date (oldest first)
    orders = await db.service_orders.find(filter_query, {"_id": 0}).sort("created_at", 1).to_list(1000)
    
    # Convert ISO strings to datetime
    for order in orders:
        if isinstance(order.get('created_at'), str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
        if isinstance(order.get('updated_at'), str):
            order['updated_at'] = datetime.fromisoformat(order['updated_at'])
    
    # Separate URGENTE orders and put them first
    urgent_orders = [o for o in orders if o.get('status') == 'URGENTE']
    normal_orders = [o for o in orders if o.get('status') != 'URGENTE']
    
    return urgent_orders + normal_orders

@api_router.get("/service-orders/stats")
async def get_service_orders_stats(current_user: User = Depends(get_current_user)):
    """Get statistics by status"""
    pipeline = [
        {
            "$group": {
                "_id": "$status",
                "count": {"$sum": 1}
            }
        }
    ]
    
    results = await db.service_orders.aggregate(pipeline).to_list(None)
    
    stats = {
        "URGENTE": 0,
        "ABERTO": 0,
        "EM ROTA": 0,
        "LIBERADO": 0,
        "PENDENCIA": 0,
        "SUSPENSO": 0,
        "DEFINIR": 0,
        "RESOLVIDO": 0
    }
    
    for result in results:
        status = result.get("_id") or "ABERTO"
        stats[status] = result.get("count", 0)
    
    stats["total"] = sum(stats.values())
    
    return stats

@api_router.get("/service-orders/export")
async def export_service_orders(current_user: User = Depends(get_current_user)):
    """Export service orders to formatted Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    # Get all orders sorted
    orders = await db.service_orders.find({}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    
    # Separate urgent orders
    urgent_orders = [o for o in orders if o.get('status') == 'URGENTE']
    normal_orders = [o for o in orders if o.get('status') != 'URGENTE']
    all_orders = urgent_orders + normal_orders
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório O.S."
    
    # Define styles
    header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
    header_font = Font(bold=True, color="000000", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["N° CHAMADO", "N° OS", "PAT", "CLIENTE", "UNIDADE", "DATA", "SITUAÇÃO"]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Add data
    for order in all_orders:
        row = [
            str(order.get('ticket_number', '')),
            str(order.get('os_number', '')),
            str(order.get('pat', '')),
            str(order.get('client_name', '')),
            str(order.get('unit', '')),
            str(order.get('opening_date', '')),
            str(order.get('status', 'ABERTO'))
        ]
        ws.append(row)
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border_style
            cell.alignment = cell_alignment
    
    # Adjust column widths
    column_widths = [15, 10, 12, 30, 20, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        iter([excel_file.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=relatorio_ordens_servico.xlsx"
        }
    )

@api_router.get("/service-orders/{order_id}", response_model=ServiceOrder)
async def get_service_order(
    order_id: str,
    current_user: User = Depends(get_current_user)
):
    order = await db.service_orders.find_one({"id": order_id}, {"_id": 0})
    
    if not order:
        raise HTTPException(status_code=404, detail="Service order not found")
    
    # Convert ISO strings to datetime
    if isinstance(order.get('created_at'), str):
        order['created_at'] = datetime.fromisoformat(order['created_at'])
    if isinstance(order.get('updated_at'), str):
        order['updated_at'] = datetime.fromisoformat(order['updated_at'])
    
    return ServiceOrder(**order)

@api_router.put("/service-orders/{order_id}", response_model=ServiceOrder)
async def update_service_order(
    order_id: str,
    order_data: ServiceOrderUpdate,
    current_user: User = Depends(get_current_user)
):
    existing_order = await db.service_orders.find_one({"id": order_id})
    
    if not existing_order:
        raise HTTPException(status_code=404, detail="Service order not found")
    
    # Update only provided fields
    update_data = {k: v for k, v in order_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.service_orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    # Get updated order
    updated_order = await db.service_orders.find_one({"id": order_id}, {"_id": 0})
    
    # Convert ISO strings to datetime
    if isinstance(updated_order.get('created_at'), str):
        updated_order['created_at'] = datetime.fromisoformat(updated_order['created_at'])
    if isinstance(updated_order.get('updated_at'), str):
        updated_order['updated_at'] = datetime.fromisoformat(updated_order['updated_at'])
    
    return ServiceOrder(**updated_order)

@api_router.delete("/service-orders/{order_id}")
async def delete_service_order(
    order_id: str,
    current_user: User = Depends(get_current_user)
):
    result = await db.service_orders.delete_one({"id": order_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Service order not found")
    
    return {"message": "Service order deleted successfully"}

# OCR route
@api_router.post("/ocr", response_model=OCRResponse)
async def process_ocr(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    try:
        # Read image file
        contents = await file.read()
        
        # Convert to base64
        image_base64 = base64.b64encode(contents).decode('utf-8')
        
        # Verify it's a valid image
        try:
            img = Image.open(BytesIO(contents))
            img.verify()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid image file")
        
        # Process OCR
        result = await extract_text_from_image(image_base64)
        
        return OCRResponse(**result)
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"OCR processing error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to process image")

# Include router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
